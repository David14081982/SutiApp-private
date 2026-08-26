#!/usr/bin/env python3
"""Apply the guarded banking schema and exact/partial historical seed."""
from __future__ import annotations
import argparse, hashlib, json, re, urllib.error, urllib.parse, urllib.request, uuid
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
SOURCE=Path(r"C:\Users\david\Downloads\Usuarios SUTIAPP.xlsx")
MIGRATION=ROOT/'supabase/migrations/20260825000200_user_maintained_historical_banking.sql'
EXPECTED_HASH='36E61B82F1BAB496B08E70BF3E1A14911A7A4E612EC3DE9F8A0669B8F2011CD3'

def require(value,message):
    if not value: raise RuntimeError(message)
def env():
    out={}
    for raw in (ROOT/'supabase.env').read_text(encoding='utf-8-sig').splitlines():
        if raw.strip() and not raw.lstrip().startswith('#') and '=' in raw:
            k,v=raw.split('=',1);out[k.strip()]=v.strip().strip('"').strip("'")
    return out
def db_query(endpoint,token,sql):
    req=urllib.request.Request(endpoint,data=json.dumps({'query':sql}).encode(),headers={'Authorization':'Bearer '+token,'Content-Type':'application/json','User-Agent':'SutiApp-BankingSeed/1.0'},method='POST')
    try:
        with urllib.request.urlopen(req,timeout=180) as res:return json.loads(res.read())
    except urllib.error.HTTPError as exc: raise RuntimeError('DATABASE_QUERY_FAILED:'+exc.read().decode('utf-8','replace')[:2000]) from exc
def rest(base,key,path,method='GET',body=None,prefer=None):
    headers={'apikey':key,'Authorization':'Bearer '+key,'Content-Type':'application/json','Accept':'application/json','User-Agent':'SutiApp-BankingSeed/1.0'}
    if prefer: headers['Prefer']=prefer
    req=urllib.request.Request(base+'/rest/v1/'+path,data=None if body is None else json.dumps(body).encode(),headers=headers,method=method)
    try:
        with urllib.request.urlopen(req,timeout=180) as res:
            payload=res.read();return json.loads(payload) if payload else []
    except urllib.error.HTTPError as exc: raise RuntimeError('REST_FAILED:'+str(exc.code)+':'+exc.read().decode('utf-8','replace')[:1000]) from exc
def text(v):
    if v is None:return None
    if isinstance(v,float) and v.is_integer():return str(int(v))
    value=str(v).strip();return value or None
def strict_digits(v):
    return re.sub(r'[ -]','',v) if v and re.fullmatch(r'[0-9 -]+',v) else None
def safe(row):
    c=strict_digits(row['clabe']);a=strict_digits(row['account']);b=row['bank'] if row['bank'] and len(row['bank'])>=2 else None
    return {'bank_name':b,'clabe':c if c and len(c)==18 else None,'account_number':a if a and 4<=len(a)<=20 else None}
def source_rows():
    require(SOURCE.is_file(),'HISTORICAL_WORKBOOK_MISSING')
    digest=hashlib.sha256(SOURCE.read_bytes()).hexdigest().upper();require(digest==EXPECTED_HASH,'SOURCE_HASH_CHANGED')
    wb=load_workbook(SOURCE,read_only=True,data_only=False);ws=wb['Usuarios'];headers=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    require(headers[0]=='Número de control' and headers[125]=='Clabe interbancaria' and headers[126]=='Número de cuenta bancario' and headers[127]=='Banco','SOURCE_HEADERS_CHANGED')
    rows=[]
    for ordinal,values in enumerate(ws.iter_rows(min_row=2,values_only=True),1):
        row={'ordinal':ordinal,'control':text(values[0]),'clabe':text(values[125]),'account':text(values[126]),'bank':text(values[127])}
        if row['clabe'] or row['account'] or row['bank']:rows.append(row)
    wb.close();require(len(rows)==513,'BANKING_EVIDENCE_COUNT_CHANGED');return digest,rows
def main():
    parser=argparse.ArgumentParser();parser.add_argument('--apply-schema',action='store_true');parser.add_argument('--apply-seed',action='store_true');args=parser.parse_args()
    values=env();base=values['SUPABASE_URL'].rstrip('/');ref=urllib.parse.urlsplit(base).hostname.split('.')[0];endpoint=f'https://api.supabase.com/v1/projects/{ref}/database/query'
    sql=MIGRATION.read_text(encoding='utf-8')
    applied=db_query(endpoint,values['SUPABASE_ACCESS_TOKEN'],"select exists(select 1 from information_schema.columns where table_schema='public' and table_name='affiliate_bank_accounts' and column_name='data_status') applied")[0]['applied']
    if not applied:
        if args.apply_schema: db_query(endpoint,values['SUPABASE_ACCESS_TOKEN'],sql)
        else:
            dry=sql.rstrip();dry=dry[:-7]+'rollback;' if dry.lower().endswith('commit;') else dry+'\nrollback;';db_query(endpoint,values['SUPABASE_ACCESS_TOKEN'],dry)
    if args.apply_seed and not (applied or args.apply_schema):raise RuntimeError('SCHEMA_NOT_APPLIED')
    digest,rows=source_rows();controls=Counter(r['control'] for r in rows if r['control'])
    key=values['SUPABASE_SECRET_KEY'];affiliates=rest(base,key,'affiliates?select=id,numero_control&limit=2000')
    existing=rest(base,key,'affiliate_bank_accounts?select=id,affiliate_id,source_file_hash,source_row_ordinal&limit=2000') if (applied or args.apply_schema) else rest(base,key,'affiliate_bank_accounts?select=id,affiliate_id&limit=2000')
    by_control=defaultdict(list)
    for a in affiliates:by_control[a['numero_control']].append(a)
    existing_affiliates={r['affiliate_id'] for r in existing};existing_source={(r.get('source_file_hash'),r.get('source_row_ordinal')) for r in existing}
    inserts=[];ambiguous=unrecoverable=conflicts=no_op=0;now=datetime.now(timezone.utc).isoformat()
    for row in rows:
        matches=by_control.get(row['control'],[])
        if not row['control'] or controls[row['control']]!=1 or len(matches)!=1:ambiguous+=1;continue
        fields=safe(row)
        if not any(fields.values()):unrecoverable+=1;continue
        affiliate_id=matches[0]['id']
        if (digest,row['ordinal']) in existing_source:no_op+=1;continue
        if affiliate_id in existing_affiliates:conflicts+=1;continue
        missing=['account_holder']+[name for name in ('bank_name','clabe','account_number') if not fields[name]]
        inserts.append({'id':str(uuid.uuid4()),'affiliate_id':affiliate_id,**fields,'is_primary':False,'data_status':'INCOMPLETE_HISTORICAL_DATA','incomplete_fields':missing,'source_kind':'HISTORICAL_SEED','source_file_hash':digest,'source_row_ordinal':row['ordinal'],'seeded_at':now})
    result={'rows_with_banking_evidence':len(rows),'potential_safe_records':len(inserts),'skipped_ambiguous':ambiguous,'skipped_unrecoverable':unrecoverable,'existing_no_op':no_op,'conflicts':conflicts,'existing_current_records_overwritten':0,'clabe_reconstructed':0,'account_numbers_reconstructed':0,'heuristic_matches':0,'writes':0}
    if args.apply_seed:
        inserted=[]
        try:
            for at in range(0,len(inserts),100):
                batch=inserts[at:at+100];rest(base,key,'affiliate_bank_accounts','POST',batch,'return=representation');inserted.extend(x['id'] for x in batch)
        except Exception:
            if inserted:rest(base,key,'affiliate_bank_accounts?id=in.('+','.join(inserted)+')','DELETE')
            raise
        result['writes']=len(inserted)
    print(json.dumps({'status':'PASS','schema_already_applied':applied,'schema_action':'APPLIED' if args.apply_schema and not applied else ('DRY_RUN' if not applied else 'NO_OP'),**result},sort_keys=True))
if __name__=='__main__':main()
