#!/usr/bin/env python3
"""Build the bounded, non-financial program catalog snapshot from SutiApp Final.

Google remains read-only provenance. Financial/request columns are retained in
source_payload but are never projected as runtime calculations or policy.
"""
from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import re
import uuid
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

NAMESPACE = uuid.UUID("46a05c64-669b-42af-a1aa-f398d66e9318")

SHEETS = {
    "1 Vehículos SutiAuto": {"program": "auto", "name": ("Marca", "Modelo", "Año"), "price": "Precio", "legacy": True},
    "2 Vehículos en renta": {"program": "renta", "name": ("Marca", "Modelo", "Año"), "price": "Precio", "legacy": True},
    "6 Suti Terrenos": {"program": "terrenos", "name": ("Título",), "description": "Observaciones", "price": "PRECIO DE CONTADO", "legacy": True},
    "7 Suti Tours": {"program": "tours", "name": ("Título",), "description": "Descripción", "category": "Categoría", "price": "Precio de contado", "legacy": True},
    "8 Suti Farma": {"program": "farma", "name": ("NOMBRE",), "category": "Categoria", "quantity": "CANTIDAD", "presentation": "GRAMOS", "contact": "URL Whatsapp", "legacy": False, "request_mode": "supabase"},
    "10 Donativos": {"program": "donativos", "name": ("Título",), "description": "Descripción", "category": "Categoría", "price": "Precio a donar", "legacy": True},
    "Suti Casa": {"program": "casa", "name": ("Título",), "description": "Descripción", "category": "Tipo de Propiedad", "price": "Precio", "contact": "Contacto", "legacy": True},
    "Paneles Solares": {"program_by_category": {"paneles solares": "solar", "aires acondicionados": "aires", "puertas de seguridad": "puertas", "equipos de computo": "computo", "equipos de cómputo": "computo"}, "name": ("Título",), "description": "Descripción", "category": "Categoría", "price": "Precio de contado", "legacy": True},
}

URL_RE = re.compile(r"https?://", re.I)


def text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def money(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    raw = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(raw) if raw else None
    except ValueError:
        return None


def json_value(value):
    if isinstance(value, (datetime.date, datetime.datetime, datetime.time)):
        return value.isoformat()
    return value


def canonical_hash(value):
    data = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(data).hexdigest().upper()


def build(source: Path):
    workbook = load_workbook(source, read_only=True, data_only=False)
    items = []
    for sheet_name, spec in SHEETS.items():
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [text(value) or f"COLUMN_{index}" for index, value in enumerate(next(rows), start=1)]
        for source_row, values in enumerate(rows, start=2):
            raw = {headers[index]: json_value(value) for index, value in enumerate(values) if index < len(headers) and value not in (None, "")}
            if not raw:
                continue
            category = text(raw.get(spec.get("category")))
            program = spec.get("program")
            if not program:
                program = spec["program_by_category"].get((category or "").casefold())
            if not program:
                continue
            parts = [text(raw.get(key)) for key in spec["name"]]
            name = " ".join(part for part in parts if part)
            if not name:
                name = category
            if not name:
                continue
            assets = []
            for index, value in enumerate(values, start=1):
                header = headers[index - 1]
                if isinstance(value, str) and URL_RE.match(value.strip()) and "video" not in header.casefold() and not (spec.get("contact") and header == spec["contact"]):
                    assets.append({"source_column": headers[index - 1], "source_column_letter": get_column_letter(index), "source_url_sha256": hashlib.sha256(value.strip().encode()).hexdigest().upper()})
            item = {
                "id": str(uuid.uuid5(NAMESPACE, f"program-catalog:{sheet_name}:{source_row}")),
                "program_key": program,
                "name": name,
                "description": text(raw.get(spec.get("description"))),
                "category_raw": category,
                "quantity_raw": text(raw.get(spec.get("quantity"))),
                "presentation_raw": text(raw.get(spec.get("presentation"))),
                "contact_url_raw": text(raw.get(spec.get("contact"))),
                "price_cash": money(raw.get(spec.get("price"))),
                "requires_quote": program not in {"farma"},
                "request_mode": spec.get("request_mode", "legacy_pending"),
                "legacy_boundary": bool(spec.get("legacy")),
                "enabled": True,
                "sort_order": source_row,
                "source_sheet": sheet_name,
                "source_row_ordinal": source_row,
                "source_payload": raw,
                "assets": assets,
            }
            items.append(item)
    core = {"source_file": "SutiApp Final", "items": items}
    return {**core, "source_snapshot_hash": canonical_hash(core), "counts": {key: sum(item["program_key"] == key for item in items) for key in sorted({item["program_key"] for item in items})}}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, default=Path("data/program-catalog-source.json"))
    args = parser.parse_args()
    result = build(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": "PASS", "items": len(result["items"]), "counts": result["counts"], "snapshot": result["source_snapshot_hash"]}, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
