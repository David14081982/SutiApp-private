#!/usr/bin/env python3
"""Read-only classifier for the owner-authorized historical banking seed.

Reads the exact Excel workbook and Supabase through SELECT-only Management API
queries. Output is aggregate/masked evidence: complete banking identifiers,
affiliate UUIDs and secrets are never printed or persisted.
"""
from __future__ import annotations

import hashlib
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\david\Downloads\Usuarios SUTIAPP.xlsx")
SHEET = "Usuarios"
CERTIFIED_HASH = "F4BA18ABE82B148ED65737DB16074303627F96D37FA6F9F025E0A10649BD9591"
EXPECTED_HEADERS = {
    1: "Número de control",
    126: "Clabe interbancaria",
    127: "Número de cuenta bancario",
    128: "Banco",
}


def require(condition: bool, message: str):
    if not condition:
        raise RuntimeError(message)


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in (ROOT / "supabase.env").read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def source_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def query(endpoint: str, token: str, sql: str):
    require(re.match(r"^\s*(select|with)\b", sql, re.I) is not None, "NON_READ_ONLY_SQL_DENIED")
    require(re.search(r"\b(insert|update|delete|merge|alter|create|drop|truncate|grant|revoke)\b", sql, re.I) is None, "WRITE_KEYWORD_DENIED")
    request = urllib.request.Request(
        endpoint,
        data=json.dumps({"query": sql}).encode(),
        headers={
            "Authorization": "Bearer " + token,
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "SutiApp-BankingSeedDryRun/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            return json.loads(response.read())
    except urllib.error.HTTPError as error:
        raise RuntimeError("DATABASE_READ_FAILED:" + error.read().decode("utf-8", "replace")[:600]) from error


def text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    value = str(value).strip()
    return value or None


def strict_digits(value: str | None) -> str | None:
    """Only remove visible grouping separators; never reconstruct E notation."""
    if not value or re.fullmatch(r"[0-9 -]+", value) is None:
        return None
    return re.sub(r"[ -]", "", value)


def normalized_bank(value: str | None) -> str | None:
    return " ".join(value.split()).casefold() if value else None


def tuple_key(clabe: str | None, account: str | None, bank: str | None):
    return (strict_digits(clabe), strict_digits(account), normalized_bank(bank))


def valid_tuple(key) -> bool:
    clabe, account, bank = key
    return bool(bank and len(bank) >= 2 and ((clabe and len(clabe) == 18) or (account and 4 <= len(account) <= 20)))


def valid_source_row(row: dict) -> bool:
    clabe = strict_digits(row["clabe"])
    account = strict_digits(row["account"])
    clabe_valid = row["clabe"] is None or (clabe is not None and len(clabe) == 18)
    account_valid = row["account"] is None or (account is not None and 4 <= len(account) <= 20)
    return bool(row["bank"] and len(row["bank"]) >= 2 and clabe_valid and account_valid and (clabe or account))


def safe_fields(row: dict) -> dict:
    """Preserve only demonstrable fields; an invalid sibling never discards them."""
    clabe = strict_digits(row["clabe"])
    account = strict_digits(row["account"])
    bank = row["bank"] if row["bank"] and len(row["bank"]) >= 2 else None
    return {
        "bank": bank,
        "clabe": clabe if clabe and len(clabe) == 18 else None,
        "account": account if account and 4 <= len(account) <= 20 else None,
    }


def mask_control(value: str | None) -> str:
    if not value:
        return "EMPTY"
    return "•" * max(2, len(value) - 4) + value[-4:]


def digest_rows(rows) -> str:
    payload = "\n".join("|".join(map(str, row)) for row in sorted(rows))
    return hashlib.sha256(payload.encode()).hexdigest().upper()


def main():
    require(SOURCE.is_file(), "HISTORICAL_WORKBOOK_MISSING")
    actual_hash = source_hash(SOURCE)
    workbook = load_workbook(SOURCE, read_only=True, data_only=False)
    require(SHEET in workbook.sheetnames, "SOURCE_SHEET_MISSING")
    sheet = workbook[SHEET]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    require(len(headers) == 187, "SOURCE_COLUMN_COUNT_CHANGED")
    for position, expected in EXPECTED_HEADERS.items():
        require(headers[position - 1] == expected, f"HEADER_MISMATCH:{position}")

    rows = []
    for ordinal, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        rows.append({
            "ordinal": ordinal,
            "control": text(values[0]),
            "clabe": text(values[125]),
            "account": text(values[126]),
            "bank": text(values[127]),
        })
    workbook.close()
    require(len(rows) == 947, "SOURCE_ROW_COUNT_CHANGED")
    banking = [row for row in rows if row["clabe"] or row["account"] or row["bank"]]
    source_controls = Counter(row["control"] for row in rows if row["control"] is not None)

    values = load_env()
    for name in ("SUPABASE_URL", "SUPABASE_ACCESS_TOKEN"):
        require(values.get(name), f"MISSING_ENV:{name}")
    ref = urllib.parse.urlsplit(values["SUPABASE_URL"]).hostname.split(".")[0]
    endpoint = f"https://api.supabase.com/v1/projects/{ref}/database/query"
    token = values["SUPABASE_ACCESS_TOKEN"]
    affiliates = query(endpoint, token, "select id::text,numero_control from public.affiliates order by id")
    current = query(endpoint, token, "select id::text,affiliate_id::text,bank_name,clabe,account_number from public.affiliate_bank_accounts order by id")
    db_by_control: dict[str, list[dict]] = defaultdict(list)
    for affiliate in affiliates:
        if affiliate["numero_control"] is not None:
            db_by_control[affiliate["numero_control"]].append(affiliate)
    current_by_affiliate: dict[str, list[dict]] = defaultdict(list)
    for account in current:
        current_by_affiliate[account["affiliate_id"]].append(account)

    exact, ambiguous, no_match = [], [], []
    potential, invalid, unrecoverable, no_op, conflicts = [], [], [], [], []
    mapping_evidence = []
    for row in banking:
        control = row["control"]
        matches = db_by_control.get(control, []) if control is not None else []
        if control is not None and source_controls[control] == 1 and len(matches) == 1:
            exact.append(row)
            affiliate = matches[0]
            key = tuple_key(row["clabe"], row["account"], row["bank"])
            mapping_evidence.append((row["ordinal"], control, affiliate["id"]))
            safe = safe_fields(row)
            if not all((safe["bank"], safe["clabe"], safe["account"])): invalid.append(row)
            if not any(safe.values()):
                unrecoverable.append(row); continue
            existing = current_by_affiliate.get(affiliate["id"], [])
            if not existing:
                potential.append(row)
                continue
            current_keys = {(item["clabe"], item["account_number"], normalized_bank(item["bank_name"])) for item in existing}
            if key in current_keys:
                no_op.append(row)
            else:
                conflicts.append(row)
        elif control is not None and (source_controls[control] > 1 or len(matches) > 1):
            ambiguous.append(row)
        else:
            no_match.append(row)

    exact_full = Counter((r["control"], r["clabe"], r["account"], r["bank"]) for r in banking)
    repeated_credentials = Counter((r["clabe"], r["account"], r["bank"]) for r in banking)
    invalid_total = [row for row in banking if not valid_source_row(row)]
    scientific_notation = [row for row in banking if re.search(r"\d(?:\.\d+)?E[+-]?\d+", (row["clabe"] or "") + " " + (row["account"] or ""), re.I)]
    per_control: dict[str, set] = defaultdict(set)
    for row in banking:
        if row["control"]:
            per_control[row["control"]].add((row["clabe"], row["account"], row["bank"]))

    result = {
        "status": "DRY_RUN_PASS",
        "source": {
            "path": str(SOURCE), "sheet": SHEET, "sha256": actual_hash,
            "certified_sha256": CERTIFIED_HASH, "certified_hash_match": actual_hash == CERTIFIED_HASH,
            "headers": {str(k): v for k, v in EXPECTED_HEADERS.items()},
        },
        "rows_scanned": len(rows),
        "rows_with_banking_data": len(banking),
        "rows_with_clabe": sum(row["clabe"] is not None for row in rows),
        "rows_with_account_number": sum(row["account"] is not None for row in rows),
        "rows_with_bank": sum(row["bank"] is not None for row in rows),
        "rows_with_all_three_fields": sum(bool(row["clabe"] and row["account"] and row["bank"]) for row in rows),
        "exact_affiliate_matches": len(exact),
        "ambiguous": len(ambiguous),
        "no_match": len(no_match),
        "duplicate_banking_rows": sum(n for n in exact_full.values() if n > 1),
        "duplicate_banking_groups": sum(n > 1 for n in exact_full.values()),
        "repeated_credential_rows_across_controls": sum(n for n in repeated_credentials.values() if n > 1),
        "affiliates_with_multiple_historical_accounts": sum(len(items) > 1 for items in per_control.values()),
        "existing_supabase_banking_records": len(current),
        "potential_inserts": len(potential),
        "potential_updates": 0,
        "existing_current_data_no_op": len(no_op),
        "conflicts": len(conflicts),
        "invalid_source_values": len(invalid_total),
        "invalid_source_values_with_exact_match": len(invalid),
        "incomplete_historical_data": len(invalid),
        "skipped_unrecoverable": len(unrecoverable),
        "rows_with_scientific_notation": len(scientific_notation),
        "mapping_digest_sha256": digest_rows(mapping_evidence),
        "potential_insert_digest_sha256": digest_rows((r["ordinal"], r["control"]) for r in potential),
        "masked_mapping_samples": [
            {"source_row_ordinal": row["ordinal"], "numero_control": mask_control(row["control"]), "classification": "POTENTIAL_INSERT"}
            for row in potential[:5]
        ],
        "writes": 0,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
