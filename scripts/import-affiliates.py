#!/usr/bin/env python3
"""Profile and import the authoritative affiliate workbook without persisting PII locally."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

EXPECTED_HASH = "F4BA18ABE82B148ED65737DB16074303627F96D37FA6F9F025E0A10649BD9591"
EXPECTED_ROWS = 947
EXPECTED_COLUMNS = 187
SHEET = "Usuarios"

LOCAL_RE = re.compile(r"^[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+$")
DOMAIN_LABEL_RE = re.compile(r"^[A-Za-z0-9-]+$")

FIELD_POSITIONS = {
    "numero_control": 1,
    "historical_status_raw": 2,
    "full_name": 3,
    "display_name": 4,
    "historical_email_raw": 5,
    "phone_raw": 9,
    "unit_raw": 11,
    "city_raw": 12,
    "employment_position_raw": 13,
    "employment_entry_date_raw": 14,
    "occupation_raw": 18,
    "institute_entry_date_raw": 19,
    "union_enrollment_date_raw": 20,
    "capture_date_raw": 21,
    "affiliation_raw": 22,
    "birth_date_raw": 23,
    "gender_raw": 24,
    "employment_area_raw": 25,
    "address_raw": 26,
    "marital_status_raw": 27,
    "children_count_raw": 28,
    "affiliate_status_raw": 29,
    "union_position_raw": 30,
    "rfc_raw": 31,
    "employment_level_raw": 32,
    "pension_raw": 116,
    "termination_date_raw": 142,
    "curp_raw": 144,
    "subdirectorate_raw": 148,
}

IMPORT_FIELDS = [
    *FIELD_POSITIONS.keys(),
    "historical_email_normalized",
    "auth_eligibility",
    "auth_ineligibility_reason",
    "source_row_ordinal",
    "source_file_hash",
]


class ImportFailure(RuntimeError):
    pass


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        name, value = line.split("=", 1)
        values[name.strip()] = value.strip().strip('"').strip("'")
    return values


def source_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def to_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalized_email(raw: str | None) -> str | None:
    if raw is None or not raw.strip():
        return None
    value = unicodedata.normalize("NFC", raw.strip()).casefold()
    if value.count("@") != 1:
        return value
    local, domain = value.split("@", 1)
    try:
        ascii_domain = domain.encode("idna").decode("ascii").lower()
    except UnicodeError:
        return value
    return f"{local}@{ascii_domain}"


def valid_email(normalized: str | None) -> bool:
    if normalized is None or normalized.count("@") != 1:
        return False
    local, domain = normalized.split("@", 1)
    if not 1 <= len(local) <= 64 or not LOCAL_RE.fullmatch(local):
        return False
    if local.startswith(".") or local.endswith(".") or ".." in local:
        return False
    if len(domain) > 253 or "." not in domain:
        return False
    labels = domain.split(".")
    return all(
        1 <= len(label) <= 63
        and DOMAIN_LABEL_RE.fullmatch(label)
        and not label.startswith("-")
        and not label.endswith("-")
        for label in labels
    )


def load_rows(path: Path) -> tuple[list[dict[str, object]], dict[str, int]]:
    actual_hash = source_hash(path)
    if actual_hash != EXPECTED_HASH:
        raise ImportFailure(f"Source hash mismatch: expected {EXPECTED_HASH}, got {actual_hash}")

    workbook = load_workbook(path, read_only=True, data_only=False)
    if SHEET not in workbook.sheetnames:
        raise ImportFailure(f"Required sheet missing: {SHEET}")
    sheet = workbook[SHEET]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if len(header) != EXPECTED_COLUMNS:
        raise ImportFailure(f"Expected {EXPECTED_COLUMNS} columns, got {len(header)}")
    if header[0] != "Número de control" or header[4] != "Email":
        raise ImportFailure("Authoritative control/email headers do not match H-003")

    rows: list[dict[str, object]] = []
    for ordinal, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        if not any(value is not None for value in values):
            raise ImportFailure(f"Unexpected fully empty source row at ordinal {ordinal}")
        row = {field: to_text(values[position - 1]) for field, position in FIELD_POSITIONS.items()}
        row["historical_email_normalized"] = normalized_email(row["historical_email_raw"])
        row["source_row_ordinal"] = ordinal
        row["source_file_hash"] = actual_hash
        rows.append(row)
    workbook.close()

    if len(rows) != EXPECTED_ROWS:
        raise ImportFailure(f"Expected {EXPECTED_ROWS} source rows, got {len(rows)}")

    valid_counts = Counter(
        row["historical_email_normalized"]
        for row in rows
        if valid_email(row["historical_email_normalized"])
    )
    seen_valid: Counter[str] = Counter()
    for row in rows:
        raw = row["historical_email_raw"]
        normalized = row["historical_email_normalized"]
        if raw is None or not str(raw).strip():
            eligibility = "missing_email"
        elif not valid_email(normalized):
            eligibility = "invalid_email"
        else:
            seen_valid[str(normalized)] += 1
            eligibility = "eligible" if seen_valid[str(normalized)] == 1 else "duplicate_email"
        row["auth_eligibility"] = eligibility
        row["auth_ineligibility_reason"] = None if eligibility == "eligible" else eligibility

    controls = [row["numero_control"] for row in rows if row["numero_control"] is not None]
    control_counts = Counter(controls)
    statuses = Counter(str(row["auth_eligibility"]) for row in rows)
    valid_total = sum(1 for row in rows if valid_email(row["historical_email_normalized"]))
    metrics = {
        "source_rows": len(rows),
        "processed_rows": len(rows),
        "empty_numero_control": len(rows) - len(controls),
        "duplicate_numero_control_groups": sum(1 for count in control_counts.values() if count > 1),
        "duplicate_numero_control_rows": sum(count for count in control_counts.values() if count > 1),
        "valid_emails": valid_total,
        "empty_emails": statuses["missing_email"],
        "invalid_emails": statuses["invalid_email"],
        "duplicate_email_groups": sum(1 for count in valid_counts.values() if count > 1),
        "duplicate_emails": statuses["duplicate_email"],
        "auth_eligible": statuses["eligible"],
    }
    expected = {
        "source_rows": 947,
        "processed_rows": 947,
        "empty_numero_control": 9,
        "duplicate_numero_control_groups": 13,
        "duplicate_numero_control_rows": 28,
        "valid_emails": 909,
        "empty_emails": 28,
        "invalid_emails": 10,
        "duplicate_email_groups": 5,
        "duplicate_emails": 5,
        "auth_eligible": 904,
    }
    if metrics != expected:
        raise ImportFailure(f"H-003 metric reconciliation failed: {json.dumps(metrics, sort_keys=True)}")
    return rows, metrics


def api_request(
    url: str,
    admin_key: str,
    method: str = "GET",
    body: bytes | None = None,
    prefer: str | None = None,
) -> tuple[int, bytes]:
    headers = {
        "apikey": admin_key,
        "Accept": "application/json",
    }
    # Modern sb_secret keys are opaque API keys, not JWTs. Legacy
    # service_role JWTs still require the Authorization header.
    if not admin_key.startswith("sb_secret_"):
        headers["Authorization"] = f"Bearer {admin_key}"
    if body is not None:
        headers["Content-Type"] = "application/json"
    if prefer:
        headers["Prefer"] = prefer
    request = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            return response.status, response.read()
    except urllib.error.HTTPError as error:
        error.read()
        raise ImportFailure(f"Supabase API request failed with HTTP {error.code}") from None
    except urllib.error.URLError as error:
        raise ImportFailure(f"Supabase API connection failed: {error.reason}") from None


def remote_rows(url: str, admin_key: str) -> list[dict[str, object]]:
    select = urllib.parse.quote(",".join(IMPORT_FIELDS), safe=",")
    query = urllib.parse.urlencode({"source_file_hash": f"eq.{EXPECTED_HASH}", "limit": "1000"})
    _, payload = api_request(f"{url.rstrip('/')}/rest/v1/affiliates?select={select}&{query}", admin_key)
    value = json.loads(payload)
    if not isinstance(value, list):
        raise ImportFailure("Unexpected Supabase reconciliation response")
    return value


def canonical(rows: list[dict[str, object]]) -> str:
    projected = [{field: row.get(field) for field in IMPORT_FIELDS} for row in rows]
    projected.sort(key=lambda row: int(row["source_row_ordinal"]))
    return hashlib.sha256(
        json.dumps(projected, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest().upper()


def apply_import(rows: list[dict[str, object]], url: str, admin_key: str) -> dict[str, object]:
    existing = remote_rows(url, admin_key)
    if len(existing) not in (0, EXPECTED_ROWS):
        raise ImportFailure(
            f"Partial destination state detected ({len(existing)}/{EXPECTED_ROWS}); no overwrite attempted"
        )
    inserted = 0
    if not existing:
        payload = json.dumps(rows, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        status, _ = api_request(
            f"{url.rstrip('/')}/rest/v1/affiliates",
            admin_key,
            method="POST",
            body=payload,
            prefer="return=minimal",
        )
        if status not in (200, 201):
            raise ImportFailure(f"Unexpected insert status: {status}")
        inserted = len(rows)
        existing = remote_rows(url, admin_key)

    if len(existing) != EXPECTED_ROWS:
        raise ImportFailure(f"Destination reconciliation failed: {len(existing)}/{EXPECTED_ROWS}")
    if canonical(rows) != canonical(existing):
        raise ImportFailure("Destination fingerprint differs from authoritative processed rows")
    return {
        "destination_rows": len(existing),
        "inserted_rows": inserted,
        "rejected_rows": 0,
        "lost_rows": 0,
        "fingerprint_match": True,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--env-file", type=Path, default=Path("supabase.env"))
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    try:
        rows, metrics = load_rows(args.source)
        result: dict[str, object] = {
            **metrics,
            "mode": "apply" if args.apply else "dry-run",
            "source_hash": EXPECTED_HASH,
            "destination_rows": None,
            "inserted_rows": 0,
            "rejected_rows": 0,
            "lost_rows": 0,
        }
        if args.apply:
            env = read_env(args.env_file)
            url = env.get("SUPABASE_URL", "")
            admin_key = env.get("SUPABASE_SECRET_KEY", "") or env.get(
                "SUPABASE_SERVICE_ROLE_KEY", ""
            )
            if not url or not admin_key:
                raise ImportFailure(
                    "Apply requires SUPABASE_URL and a server-only SUPABASE_SECRET_KEY"
                )
            result.update(apply_import(rows, url, admin_key))
        print(json.dumps(result, ensure_ascii=False, sort_keys=True))
        return 0
    except ImportFailure as error:
        print(f"IMPORT FAILED: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
