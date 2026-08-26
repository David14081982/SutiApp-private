#!/usr/bin/env python3
"""Inventory historical file URLs without modifying source workbooks.

Raw URLs and row-level coordinates are written only to an explicitly supplied
private output path (normally C:\\tmp).  The repository catalog contains only
column semantics and aggregate counts.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPECTED_USERS_HASH = "F4BA18ABE82B148ED65737DB16074303627F96D37FA6F9F025E0A10649BD9591"
URL_RE = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)
TRAILING = ").,;]}"
FILE_EXTENSIONS = {
    ".jpg", ".jpeg", ".jfif", ".png", ".webp", ".gif", ".svg", ".pdf", ".heic",
    ".heif", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".ppt", ".pptx", ".txt",
    ".zip", ".rar",
}

PRIVATE_HINTS = (
    "ine", "identificacion", "identificación", "talon", "talón", "nomina", "nómina",
    "comprobante", "domicilio", "credencial", "firma", "documento personal",
    "foto perfil", "foto de perfil", "avatar", "solicitud", "estado de cuenta",
)
PUBLIC_HINTS = (
    "banner", "logo", "portada", "imagen institucional", "reglamento", "norma",
    "minuta", "formato", "tutorial", "educacion", "educación", "publicidad",
)
FINANCIAL_HINTS = (
    "ahorro", "prestamo", "préstamo", "amortizacion", "amortización", "fondos",
    "conciliacion", "conciliación", "nomina", "nómina", "sutiauto", "inversion",
    "inversión", "adelanto", "solicitudes", "historial p", "reporte ahorro",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def normalized(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).casefold()


def canonical_url(raw: str) -> str:
    value = raw.rstrip(TRAILING)
    parsed = urlsplit(value)
    host = (parsed.hostname or "").lower()
    port = f":{parsed.port}" if parsed.port else ""
    netloc = host + port
    return urlunsplit((parsed.scheme.lower(), netloc, parsed.path, parsed.query, ""))


def urls_from(value: object) -> list[str]:
    if not isinstance(value, str) or "http" not in value.casefold():
        return []
    found: list[str] = []
    for match in URL_RE.findall(value):
        try:
            url = canonical_url(match)
            if urlsplit(url).scheme in {"http", "https"} and urlsplit(url).hostname:
                found.append(url)
        except ValueError:
            continue
    return found


def is_file_candidate(url: str) -> bool:
    parsed = urlsplit(url)
    suffix = Path(parsed.path).suffix.casefold()
    if suffix in FILE_EXTENSIONS:
        return True
    host = (parsed.hostname or "").casefold()
    path = parsed.path.casefold()
    return (
        host == "storage.googleapis.com" and "glide-prod.appspot.com/" in path
    ) or host == "firebasestorage.googleapis.com"


def semantic_file_key(semantic: str) -> str:
    key = normalized(semantic)
    known = (
        (("ine frente", "ine frontal"), "ine_front"),
        (("ine reverso",), "ine_back"),
        (("talon ultima", "ultimo talon", "ultima quincena"), "payroll_receipt_latest"),
        (("talon penultima", "penultima quincena"), "payroll_receipt_previous"),
        (("foto perfil", "foto de perfil", "fotografia rostro", "imagen de perfil", "photo"), "profile_photo"),
        (("comprobante de domicilio",), "address_proof"),
        (("hoja de afiliacion",), "membership_form"),
        (("hoja tribunal",), "tribunal_form"),
        (("credencial afiliado",), "credential"),
        (("firma",), "signature"),
        (("constancia de no adeudo",), "no_debt_certificate"),
        (("pdf",), "pdf_document"),
    )
    for hints, result in known:
        if any(hint in key for hint in hints):
            return result
    slug = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    return slug[:80] or "other_document"


def classify(source_system: str, sheet: str, semantic: str) -> tuple[str, str, str, str]:
    combined = normalized(f"{sheet} {semantic}")
    if source_system == "owner_master_excel":
        return "PRIVATE", "affiliate", "affiliate_files", "AFFILIATE_EXACT_MATCH_REQUIRED"
    if any(h in combined for h in FINANCIAL_HINTS):
        return "PRIVATE", "financial_legacy", "historical_asset_sources", "PENDING_DOMAIN_LINK"
    target = "historical_asset_sources"
    domain = "unresolved"
    routes = (
        (("directorio",), "directory", "directory_members"),
        (("minutas",), "minute", "minutes"),
        (("descarga", "normas", "reglamentos"), "institutional_document", "institutional_documents"),
        (("anuncio principal", "banner"), "banner", "banners"),
        (("promociones",), "popup", "popups"),
        (("convenios2",), "company", "company_assets"),
        (("informacion educativa", "tutoriales"), "education", "educational_resources"),
        (("secretaria de finanzas",), "institutional_program", "institutional_programs"),
        (("membresias",), "membership", "membership_offerings"),
        (("categorias suticompras", "productos suticompras"), "marketplace", "marketplace_product_assets"),
    )
    for hints, route_domain, relation in routes:
        if any(h in combined for h in hints):
            domain, target = route_domain, relation
            break
    if any(h in combined for h in PRIVATE_HINTS):
        return "PRIVATE", domain, target, "PENDING_OWNER_REVIEW"
    if any(h in combined for h in PUBLIC_HINTS) or domain in {
        "directory", "minute", "institutional_document", "banner", "popup", "company",
        "education", "institutional_program", "membership", "marketplace",
    }:
        return "PUBLIC", domain, target, "DOMAIN_LINK_CANDIDATE"
    return "PRIVATE", domain, target, "PENDING_SECURITY_CLASSIFICATION"


def header_map(sheet) -> dict[int, str]:
    headers: dict[int, str] = {}
    for column, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1), ()), start=1):
        value = "" if cell.value is None else str(cell.value).strip()
        headers[column] = value or get_column_letter(column)
    return headers


def find_control_column(headers: dict[int, str]) -> int | None:
    candidates = [column for column, value in headers.items() if normalized(value) in {
        "numero de control", "numero control", "no. de control", "no de control", "numero_control"
    }]
    return candidates[0] if len(candidates) == 1 else None


def scan_workbook(path: Path, source_system: str, source_file: str) -> tuple[list[dict], list[dict], dict]:
    workbook_hash = sha256(path)
    if source_system == "owner_master_excel" and workbook_hash != EXPECTED_USERS_HASH:
        raise RuntimeError(f"Usuarios SUTIAPP.xlsx hash mismatch: {workbook_hash}")
    wb = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    records: list[dict] = []
    columns: dict[tuple[str, str], dict] = {}
    metrics = {"sheets": len(wb.sheetnames), "rows_scanned": 0, "cells_scanned": 0}
    try:
        for ws in wb.worksheets:
            headers = header_map(ws)
            control_column = find_control_column(headers)
            for row_number, row in enumerate(ws.iter_rows(), start=1):
                metrics["rows_scanned"] += 1
                for column, cell in enumerate(row, start=1):
                    metrics["cells_scanned"] += 1
                    extracted_all = urls_from(cell.value)
                    hyperlink = getattr(cell, "hyperlink", None)
                    if hyperlink is not None:
                        extracted_all += urls_from(getattr(hyperlink, "target", None))
                    extracted = [url for url in extracted_all if is_file_candidate(url)]
                    if not extracted:
                        continue
                    semantic = headers.get(column, get_column_letter(column))
                    visibility, domain, relation, ownership = classify(source_system, ws.title, semantic)
                    raw_control = None
                    if control_column and row_number > 1 and control_column <= len(row):
                        value = row[control_column - 1].value
                        raw_control = None if value is None else str(value)
                    for order, url in enumerate(dict.fromkeys(extracted), start=1):
                        parsed = urlsplit(url)
                        records.append({
                            "source_system": source_system,
                            "source_file": source_file,
                            "source_file_hash": workbook_hash,
                            "source_sheet": ws.title,
                            "source_row": row_number,
                            "source_column": semantic,
                            "source_column_letter": get_column_letter(column),
                            "file_key": semantic_file_key(semantic),
                            "source_url": url,
                            "source_url_sha256": hashlib.sha256(url.encode("utf-8")).hexdigest().upper(),
                            "url_order": order,
                            "domain": parsed.hostname.lower() if parsed.hostname else None,
                            "expected_owner": "affiliate" if source_system == "owner_master_excel" else domain,
                            "classification": visibility,
                            "target_domain": domain,
                            "target_relation": relation,
                            "ownership_status": ownership,
                            "numero_control_raw": raw_control,
                        })
                    key = (ws.title, get_column_letter(column))
                    item = columns.setdefault(key, {
                        "source_system": source_system,
                        "source_file": source_file,
                        "source_file_hash": workbook_hash,
                        "sheet": ws.title,
                        "column": get_column_letter(column),
                        "semantic_name": semantic,
                        "rows_with_files": set(),
                        "urls_parsed": 0,
                        "classification": visibility,
                        "target_domain": domain,
                        "target_relation": relation,
                        "ownership_status": ownership,
                    })
                    item["rows_with_files"].add(row_number)
                    item["urls_parsed"] += len(dict.fromkeys(extracted))
    finally:
        wb.close()
    public_columns: list[dict] = []
    for item in columns.values():
        item["rows_with_files"] = len(item["rows_with_files"])
        public_columns.append(item)
    public_columns.sort(key=lambda row: (row["source_system"], row["sheet"], row["column"]))
    metrics["file_hash"] = workbook_hash
    return records, public_columns, metrics


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--users", type=Path, required=True)
    parser.add_argument("--sutiapp", type=Path, required=True)
    parser.add_argument("--raw-output", type=Path, required=True)
    parser.add_argument("--column-output", type=Path, default=ROOT / "data/master-asset-column-catalog.json")
    args = parser.parse_args()

    all_records: list[dict] = []
    all_columns: list[dict] = []
    source_metrics: dict[str, dict] = {}
    for path, system, name in (
        (args.users, "owner_master_excel", "Usuarios SUTIAPP.xlsx"),
        (args.sutiapp, "google_sheets_export", "SutiApp Final"),
    ):
        records, columns, metrics = scan_workbook(path, system, name)
        all_records.extend(records)
        all_columns.extend(columns)
        source_metrics[system] = metrics

    raw_payload = {
        "schema_version": 1,
        "records": all_records,
    }
    args.raw_output.parent.mkdir(parents=True, exist_ok=True)
    args.raw_output.write_text(json.dumps(raw_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    raw_hash = sha256(args.raw_output)

    urls = [row["source_url"] for row in all_records]
    counts = Counter(row["classification"] for row in all_records)
    domains = Counter(row["domain"] for row in all_records)
    catalog = {
        "schema_version": 1,
        "raw_inventory_sha256": raw_hash,
        "raw_inventory_location": "PRIVATE_LOCAL_OUTPUT_NOT_VERSIONED",
        "sources": source_metrics,
        "summary": {
            "historical_file_columns_discovered": len(all_columns),
            "usuarios_sutiapp_file_columns": sum(1 for row in all_columns if row["source_system"] == "owner_master_excel"),
            "sutiapp_final_file_columns": sum(1 for row in all_columns if row["source_system"] == "google_sheets_export"),
            "total_urls_discovered": len(urls),
            "unique_source_urls": len(set(urls)),
            "public_references": counts["PUBLIC"],
            "private_references": counts["PRIVATE"],
            "domains": dict(sorted(domains.items())),
        },
        "columns": all_columns,
    }
    args.column_output.parent.mkdir(parents=True, exist_ok=True)
    args.column_output.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(catalog["summary"], ensure_ascii=False, sort_keys=True))
    print(json.dumps({"raw_inventory_sha256": raw_hash, "column_catalog": str(args.column_output)}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
