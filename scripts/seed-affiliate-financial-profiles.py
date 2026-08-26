#!/usr/bin/env python3
"""Seed current financial profile fields once from the exact authorized workbook.

Dry-run is the default. --apply calls one transactional service-role RPC that
snapshots all 947 rows, validates UUID + source ordinal, audits and updates atomically.
"""
from __future__ import annotations
import argparse, hashlib, json, os, sys, unicodedata, urllib.request, uuid
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

EXPECTED_HASH="F4BA18ABE82B148ED65737DB16074303627F96D37FA6F9F025E0A10649BD9591"
EXPECTED_ROWS=947
SHEET="Usuarios"
CATEGORY_COLUMN=58
UNION_COLUMN=60

def norm(value):
    text=unicodedata.normalize("NFD",str(value or "").strip()).encode("ascii","ignore").decode().upper()
    return "_".join(text.replace("."," ").replace("/"," ").split())

CATEGORIES={norm(label):code for code,label in [
    ("SUPLENTES_VARIABLES","Suplentes Variables"),("SUPLENTES_FIJOS","Suplentes Fijos"),
    ("EVENTUALES","Eventuales"),("BASE","Base"),("JUBILADOS_PENSIONADOS","Jubilados y Pens."),("CONFIANZA","Confianza")
]}
UNIONS={norm(label):code for code,label in [
    ("SUTISSSTESON","SUTISSSTESON"),("SUEISSSTESON","SUEISSSTESON"),("SITISSSTESON","SITISSSTESON"),
    ("EMPLEADOS_DE_CONFIANZA","EMPLEADOS DE CONFIANZA"),("EXTERNO","Externo")
]}
SOURCE_ERROR_VALUES={"#N/A","#VALUE!","#REF!","#NAME?","#DIV/0!","#NULL!","#NUM!"}

def sha256(path):
    digest=hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda:handle.read(1024*1024),b""): digest.update(block)
    return digest.hexdigest().upper()

def request(url,key,method="GET",payload=None):
    data=None if payload is None else json.dumps(payload).encode()
    req=urllib.request.Request(url,data=data,method=method,headers={"apikey":key,"Authorization":f"Bearer {key}","Content-Type":"application/json","Prefer":"return=representation"})
    with urllib.request.urlopen(req,timeout=30) as response:
        return json.loads(response.read() or b"[]")

def local_env(path):
    out={}
    if not path.exists():return out
    for raw in path.read_text(encoding='utf-8-sig').splitlines():
        if raw.strip() and not raw.lstrip().startswith('#') and '=' in raw:
            key,value=raw.split('=',1);out[key.strip()]=value.strip().strip('"').strip("'")
    return out

def main():
    settings=local_env(Path(__file__).resolve().parents[1]/'supabase.env')
    parser=argparse.ArgumentParser();parser.add_argument("workbook",type=Path);parser.add_argument("--apply",action="store_true")
    parser.add_argument("--url",default=os.getenv("SUPABASE_URL","") or settings.get('SUPABASE_URL',''));parser.add_argument("--secret",default=os.getenv("SUPABASE_SERVICE_ROLE_KEY","") or os.getenv("SUPABASE_SECRET_KEY","") or settings.get('SUPABASE_SERVICE_ROLE_KEY','') or settings.get('SUPABASE_SECRET_KEY',''))
    args=parser.parse_args()
    actual=sha256(args.workbook)
    if actual!=EXPECTED_HASH: raise SystemExit(f"BLOCKED: workbook hash {actual} does not match {EXPECTED_HASH}")
    book=load_workbook(args.workbook,read_only=True,data_only=False)
    if SHEET not in book.sheetnames: raise SystemExit("BLOCKED: Usuarios sheet missing")
    sheet=book[SHEET];headers=[str(v or "").strip() for v in next(sheet.iter_rows(min_row=1,max_row=1,values_only=True))]
    if "CATEGORIA" not in norm(headers[CATEGORY_COLUMN-1]) or "SINDICATO" not in norm(headers[UNION_COLUMN-1]):
        raise SystemExit(f"BLOCKED: columns 58/60 do not match authorized headers: {headers[CATEGORY_COLUMN-1]!r}, {headers[UNION_COLUMN-1]!r}")
    rows=[];unknown=[];source_errors=0
    for ordinal,values in enumerate(sheet.iter_rows(min_row=2,values_only=True),start=1):
        category_raw=values[CATEGORY_COLUMN-1];union_raw=values[UNION_COLUMN-1]
        category_text=str(category_raw or '').strip();union_text=str(union_raw or '').strip()
        source_errors+=int(category_text in SOURCE_ERROR_VALUES)+int(union_text in SOURCE_ERROR_VALUES)
        category=CATEGORIES.get(norm(category_raw)) if category_text and category_text not in SOURCE_ERROR_VALUES else None
        union=UNIONS.get(norm(union_raw)) if union_text and union_text not in SOURCE_ERROR_VALUES else None
        if (category_text and category_text not in SOURCE_ERROR_VALUES and not category) or (union_text and union_text not in SOURCE_ERROR_VALUES and not union): unknown.append({"ordinal":ordinal,"category":category_text,"union":union_text})
        rows.append((ordinal,category,union))
    if len(rows)!=EXPECTED_ROWS: raise SystemExit(f"BLOCKED: expected {EXPECTED_ROWS} rows, got {len(rows)}")
    if unknown: raise SystemExit("BLOCKED: unknown category/union values: "+json.dumps(unknown[:20],ensure_ascii=False))
    print(json.dumps({"status":"VALIDATED","hash":actual,"rows":len(rows),"category_column":CATEGORY_COLUMN,"union_column":UNION_COLUMN,"null_categories":sum(1 for _,c,_ in rows if c is None),"null_unions":sum(1 for _,_,u in rows if u is None),"source_error_values_preserved_as_null":source_errors,"apply":args.apply}))
    if not args.apply:return 0
    if not args.url or not args.secret: raise SystemExit("BLOCKED: --apply requires SUPABASE_URL and server secret")
    rest_base=args.url.rstrip("/")+"/rest/v1/"
    current=request(rest_base+"affiliates?select=id,source_row_ordinal,source_file_hash,financial_profile_version,financial_union_code,financial_employee_category_code,financial_profile_seed_source_hash,financial_profile_seed_row_ordinal&source_file_hash=eq."+EXPECTED_HASH+"&limit=1000",args.secret)
    if len(current)!=EXPECTED_ROWS: raise SystemExit(f"BLOCKED: Supabase source universe is {len(current)}, expected {EXPECTED_ROWS}")
    if len({row['id'] for row in current})!=EXPECTED_ROWS or len({row['source_row_ordinal'] for row in current})!=EXPECTED_ROWS:
        raise SystemExit('BLOCKED: Supabase affiliate UUID/ordinal mapping is not 1:1')
    by_ordinal={row["source_row_ordinal"]:row for row in current};payload=[]
    for ordinal,category,union in rows:
        target=by_ordinal.get(ordinal)
        if not target: raise SystemExit(f"BLOCKED: missing Supabase ordinal {ordinal}")
        payload.append({'affiliate_id':target['id'],'source_row_ordinal':ordinal,'financial_employee_category_code':category,'financial_union_code':union})
    batch_id=str(uuid.uuid5(uuid.NAMESPACE_URL,'sutiapp:BULK_INITIAL_FINANCIAL_PROFILE_SEED:'+EXPECTED_HASH))
    result=request(rest_base+'rpc/bulk_seed_affiliate_financial_profiles',args.secret,'POST',{'p_batch_id':batch_id,'p_source_hash':EXPECTED_HASH,'p_rows':payload})
    if not isinstance(result,dict) or result.get('affiliates_total')!=EXPECTED_ROWS or result.get('mapping_mismatches')!=0:
        raise SystemExit('BLOCKED: transactional seed returned an invalid reconciliation contract')
    print(json.dumps({'status':'APPLIED','batch_id':batch_id,**result},sort_keys=True))
    return 0

if __name__=="__main__":sys.exit(main())
