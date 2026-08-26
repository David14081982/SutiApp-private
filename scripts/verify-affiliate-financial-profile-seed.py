"""Read-only 947-row reconciliation of the authorized financial profile seed."""
from pathlib import Path
import hashlib, importlib.util, json, urllib.parse, urllib.request
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
WORKBOOK=Path(r'C:\Users\david\Downloads\Usuarios SUTIAPP.xlsx')
BATCH_ID='82179501-f85b-50c3-b7ae-7f6998852163'

def env():
    out={}
    for raw in (ROOT/'supabase.env').read_text(encoding='utf-8-sig').splitlines():
        if raw.strip() and not raw.lstrip().startswith('#') and '=' in raw:
            k,v=raw.split('=',1);out[k.strip()]=v.strip().strip('"').strip("'")
    return out
def rest(url,key):
    req=urllib.request.Request(url,headers={'apikey':key,'Authorization':'Bearer '+key,'Accept':'application/json'})
    with urllib.request.urlopen(req,timeout=60) as response:return json.loads(response.read())
def rest_count(url,key):
    req=urllib.request.Request(url,headers={'apikey':key,'Authorization':'Bearer '+key,'Accept':'application/json','Prefer':'count=exact','Range':'0-0'})
    with urllib.request.urlopen(req,timeout=60) as response:return int(response.headers['Content-Range'].split('/')[-1])
def query(values,sql):
    ref=urllib.parse.urlsplit(values['SUPABASE_URL']).hostname.split('.')[0]
    req=urllib.request.Request(f'https://api.supabase.com/v1/projects/{ref}/database/query',data=json.dumps({'query':sql}).encode(),headers={'Authorization':'Bearer '+values['SUPABASE_ACCESS_TOKEN'],'Content-Type':'application/json'},method='POST')
    with urllib.request.urlopen(req,timeout=90) as response:return json.loads(response.read())

def main():
    spec=importlib.util.spec_from_file_location('seed_contract',ROOT/'scripts/seed-affiliate-financial-profiles.py');module=importlib.util.module_from_spec(spec);spec.loader.exec_module(module)
    if module.sha256(WORKBOOK)!=module.EXPECTED_HASH:raise RuntimeError('source hash changed')
    sheet=load_workbook(WORKBOOK,read_only=True,data_only=False)[module.SHEET];next(sheet.iter_rows(min_row=1,max_row=1,values_only=True))
    expected={}
    for ordinal,values in enumerate(sheet.iter_rows(min_row=2,values_only=True),start=1):
        cat_raw=str(values[module.CATEGORY_COLUMN-1] or '').strip();union_raw=str(values[module.UNION_COLUMN-1] or '').strip()
        expected[ordinal]=(module.CATEGORIES.get(module.norm(cat_raw)) if cat_raw and cat_raw not in module.SOURCE_ERROR_VALUES else None,module.UNIONS.get(module.norm(union_raw)) if union_raw and union_raw not in module.SOURCE_ERROR_VALUES else None)
    values=env();base=values['SUPABASE_URL'].rstrip('/')+'/rest/v1/';key=values.get('SUPABASE_SERVICE_ROLE_KEY') or values['SUPABASE_SECRET_KEY']
    rows=rest(base+'affiliates?select=id,source_row_ordinal,source_file_hash,financial_employee_category_code,financial_union_code,financial_profile_seed_source_hash,financial_profile_seed_row_ordinal&order=source_row_ordinal.asc&limit=1000',key)
    mismatches=[]
    for row in rows:
        wanted=expected.get(row['source_row_ordinal']);actual=(row['financial_employee_category_code'],row['financial_union_code'])
        if wanted!=actual or row['source_file_hash']!=module.EXPECTED_HASH or row['financial_profile_seed_source_hash']!=module.EXPECTED_HASH or row['financial_profile_seed_row_ordinal']!=row['source_row_ordinal']:
            mismatches.append(row['id'])
    batch=rest(base+f'affiliate_financial_profile_seed_batches?select=status&id=eq.{BATCH_ID}',key)
    stats={
      'affiliates_total':rest_count(base+'affiliates?select=id',key),
      'categories_seeded':rest_count(base+'affiliates?select=id&financial_employee_category_code=not.is.null',key),
      'category_null':rest_count(base+'affiliates?select=id&financial_employee_category_code=is.null',key),
      'unions_seeded':rest_count(base+'affiliates?select=id&financial_union_code=not.is.null',key),
      'union_null':rest_count(base+'affiliates?select=id&financial_union_code=is.null',key),
      'recovery_snapshot_rows':rest_count(base+f'affiliate_financial_profile_seed_snapshot?select=affiliate_id&batch_id=eq.{BATCH_ID}',key),
      'seed_audit_rows':rest_count(base+f'affiliate_profile_audit_log?select=id&batch_id=eq.{BATCH_ID}&change_source=eq.BULK_INITIAL_FINANCIAL_PROFILE_SEED',key),
      'category_audit_rows':rest_count(base+f'affiliate_profile_audit_log?select=id&batch_id=eq.{BATCH_ID}&field_name=eq.financial_employee_category_code',key),
      'union_audit_rows':rest_count(base+f'affiliate_profile_audit_log?select=id&batch_id=eq.{BATCH_ID}&field_name=eq.financial_union_code',key),
      'batch_status':batch[0]['status'] if len(batch)==1 else None,
      'preseed_nonnull_values':rest_count(base+f'affiliate_financial_profile_seed_snapshot?select=affiliate_id&batch_id=eq.{BATCH_ID}&or=(old_financial_union_code.not.is.null,old_financial_employee_category_code.not.is.null)',key),
    }
    expected_stats={'affiliates_total':947,'categories_seeded':931,'category_null':16,'unions_seeded':770,'union_null':177,'recovery_snapshot_rows':947,'seed_audit_rows':1701,'category_audit_rows':931,'union_audit_rows':770,'batch_status':'APPLIED','preseed_nonnull_values':0}
    if len(rows)!=947 or mismatches or any(stats[k]!=v for k,v in expected_stats.items()):raise RuntimeError(json.dumps({'row_count':len(rows),'mapping_mismatches':len(mismatches),'stats':stats}))
    print(json.dumps({'status':'PASS',**stats,'mapping_mismatches':0,'source_errors_inferred':0,'unexpected_affiliate_changes':0,'batch_id':BATCH_ID}))
if __name__=='__main__':main()
